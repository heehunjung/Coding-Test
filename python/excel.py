from openpyxl import load_workbook # 액셀 읽는 패키지 설치

file_path = "./이수교과목_2019202069.xlsx"

# 파일이 존재하는지 확인
if not os.path.exists(file_path):
    # 파일이 없으면 새 워크북 생성
    wb = Workbook()
    ws = wb.active
    ws.title = "성적이수표"
    wb.save(file_path)
else:
    # 파일이 있으면 로드
    wb = load_workbook(file_path)
ws = wb.active # 워크시트의 객체를 만든다. -> 기본값은 첫번째 시트가 대상. 삼성의 성적이수표 액셀파일의 첫번째 시트가 성적이수표이다.


test = [] # 중간에 값을 볼 수 있도록 테스트용 리스트를 선언

# 액셀은 0부터가 아닌 1부터 시작이다.
for i in range(1, ws.max_row + 1): # 시트의 첫번째 행 ~ 마지막 행
    for j in range(1, ws.max_column + 1): # 시트의 첫번째 열 ~ 마지막 열
        test.append(ws.cell(row=i, column = j).value) # 테스트 리스트에 액셀의 행의 값을 차례로 넣어본다.
#print(test)
fi = open("C:\\Users\\jungheehun\\Desktop\\scote.txt", "r")

gradeList = [] # 이수 정보를 담을 리스트
for i in fi:
    gradeList.append(list(i.strip().split(' '))) # 리스트에 한줄씩 삽입한다. 이 때 공백으로 분리한다.

# grade list의 한 행 안의 구성요소 
# 과목유형, 취득학점, 성적, 과목명, 수강연도, 학기
print(gradeList)

fi.close() # 텍스트 파일의 이용이 끝났으니 닫아준다.

test = [] # 테스트 리스트 초기화

# 액셀의 구성요소
# NO, 전공명, 수강연도, 학기, 과목명, 과목유형, 취득학점, 성적, 재수강여부

for i in range(len(gradeList)):
    ws.cell(row = i + 2, column = 1, value = i + 1) # 액셀의 첫번째 필드 NO -> 인덱스 i 값에 1 더해서 넣어주면 된다.
    ws.cell(row = i + 2, column = 2, value = "[기타]OO학과(OO대-학사-주전공)") # 전공명은 다음과 같은 양식으로 입력
    ws.cell(row = i + 2, column = 3, value = int(gradeList[i][4])) # 수강연도, 숫자(int)로 바꿔서 기입해준다.
    
    if(gradeList[i][5] != "동" and gradeList[i][5] != "하"):
        ws.cell(row = i + 2, column = 4, value = int(gradeList[i][5])) # 동계나 하계가 아니면 1,2 학기이므로 숫자로 변환해서 입력
    else:
        if(gradeList[i][5] == "겨"): # 동계일 경우 "겨울계절" 로 기입. 이렇게 해야 홈페이지에 업로드 했을 때 인식한다.
            ws.cell(row = i + 2, column = 4, value = "겨울계절")
        elif(gradeList[i][5] == "여"): # 하계일 경우 마찬가지
            ws.cell(row = i + 2, column = 4, value = "여름계절")
    
    # 과목 유형 분류한다. 앞에 1이 붙어서 1이 붙으면 전공으로 처리하였다. (예: 1전선, 1전심)
    # 학교마다 유형이 다르므로 텍스트 파일을 가공할 때 미리 처리하거나 여기서 값을 수정하자.
    # 값은 "전공" / "교양기타" 로 해야 자동 인식이 된다.
    if(gradeList[i][0][:1] == "1"): 
        ws.cell(row = i + 2, column = 6, value = "전공")
    else:
        ws.cell(row = i + 2, column = 6, value = "교양기타")
    
    # 취득학점과 성적을 입력한다. 숫자가 들어가야할 칸은 항상 int로 감싸서 숫자로 변환해준다.
    ws.cell(row = i + 2, column = 7, value = int(gradeList[i][1][:1]))
    ws.cell(row = i + 2, column = 8, value = gradeList[i][2])
    
    # 우리 학교는 재수강을 했을 경우, 과목명 앞에 대문자 R을 붙인다.
    # 각자 방식에 따라 수정해주자.
    # 재수강 여부 입력이다. 과목명 앞 1자리가 "R" 이면 "Y" 를 입력, 아니면 "N" 을 입력한다.
    if(gradeList[i][3][0] == "R"):
        ws.cell(row = i + 2, column = 5, value = gradeList[i][3][1:])
        ws.cell(row = i + 2, column = 9, value = "Y")
    else:
        ws.cell(row = i + 2, column = 5, value = gradeList[i][3])
        ws.cell(row = i + 2, column = 9, value = "N")

#올바로 입력했는지 테스트 리스트에 액셀 값을 넣어 테스트
for i in range(1, ws.max_row + 1):
    for j in range(1, ws.max_column + 1):
        test.append(ws.cell(row=i, column = j).value)
#print(test)

# 이수교과목.xlsx 라는 액셀파일을 새로이 만들어 여기 저장한다. 자동으로 만들어지니 미리 만들 필요가 없다.
wb.save('./이수교과목.xlsx')